import os
from division_list import get_division_list
import openpyxl


def get_xlsx_trees():
    file_list = os.listdir(r'C:\Users\saone\Documents\Python Stuff\prod\single_elimination_tree_generator\raw_divisions')

    for file in file_list:
        n, division_list, complete_tree_list, next_round_list = get_division_list(file)
        dot_index = file.index('.')
        file_name_adj = file[0:dot_index]

        if n == 2:
            wb = openpyxl.load_workbook("templates\\2.xlsx")    
            sheet = wb.active
            sheet['A1'] = f"Division: {file_name_adj}"
            
            sheet['D10'] = division_list[0][0]
            sheet['D11'] = division_list[0][1]

            sheet['D23'] = division_list[1][0]
            sheet['D24'] = division_list[1][1]
            
            wb.save(filename = f"refined_divisions\\{file_name_adj}.xlsx")     
        elif n == 4:
            wb = openpyxl.load_workbook("templates\\4.xlsx")    
            sheet = wb.active
            sheet['A1'] = f"Division: {file_name_adj}"
            
            sheet['C10'] = division_list[0][0]
            sheet['C11'] = division_list[0][1]

            sheet['C16'] = division_list[1][0]
            sheet['C17'] = division_list[1][1]
            
            sheet['C20'] = division_list[2][0]
            sheet['C21'] = division_list[2][1]

            sheet['C26'] = division_list[3][0]
            sheet['C27'] = division_list[3][1]
            
            wb.save(filename = f"refined_divisions\\{file_name_adj}.xlsx") 
        elif n == 8:
            wb = openpyxl.load_workbook("templates\\8.xlsx")    
            sheet = wb.active
            sheet['A1'] = f"Division: {file_name_adj}"
            
            sheet['B8'] = division_list[0][0]
            sheet['B9'] = division_list[0][1]
            
            sheet['B12'] = division_list[1][0]
            sheet['B13'] = division_list[1][1]
            
            sheet['B16'] = division_list[2][0]
            sheet['B17'] = division_list[2][1]
            
            sheet['B20'] = division_list[3][0]
            sheet['B21'] = division_list[3][1]
            
            sheet['B24'] = division_list[4][0]
            sheet['B25'] = division_list[4][1]
            
            sheet['B28'] = division_list[5][0]
            sheet['B29'] = division_list[5][1]
            
            sheet['B32'] = division_list[6][0]
            sheet['B33'] = division_list[6][1]
            
            sheet['B36'] = division_list[7][0]
            sheet['B37'] = division_list[7][1]
            
            wb.save(filename = f"refined_divisions\\{file_name_adj}.xlsx") 
        elif n == 16:
            wb = openpyxl.load_workbook("templates\\16.xlsx")    
            sheet = wb.active
            sheet['K6'] = f"Division: {file_name_adj}"
            
            sheet['B3'] = division_list[0][0]
            sheet['B4'] = division_list[0][1]
            
            sheet['B7'] = division_list[1][0]
            sheet['B8'] = division_list[1][1]
            
            sheet['B11'] = division_list[2][0]
            sheet['B12'] = division_list[2][1]
            
            sheet['B15'] = division_list[3][0]
            sheet['B16'] = division_list[3][1]
            
            sheet['B19'] = division_list[4][0]
            sheet['B20'] = division_list[4][1]
            
            sheet['B23'] = division_list[5][0]
            sheet['B24'] = division_list[5][1]
            
            sheet['B27'] = division_list[6][0]
            sheet['B28'] = division_list[6][1]
            
            sheet['B31'] = division_list[7][0]
            sheet['B32'] = division_list[7][1]
            
            sheet['B35'] = division_list[8][0]
            sheet['B36'] = division_list[8][1]
            
            sheet['B39'] = division_list[9][0]
            sheet['B40'] = division_list[9][1]
            
            sheet['B43'] = division_list[10][0]
            sheet['B44'] = division_list[10][1]
            
            sheet['B47'] = division_list[11][0]
            sheet['B48'] = division_list[11][1]
            
            sheet['B51'] = division_list[12][0]
            sheet['B52'] = division_list[12][1]
            
            sheet['B55'] = division_list[13][0]
            sheet['B56'] = division_list[13][1]
            
            sheet['B59'] = division_list[14][0]
            sheet['B60'] = division_list[14][1]
            
            sheet['B63'] = division_list[15][0]
            sheet['B64'] = division_list[15][1]
            
            wb.save(filename = f"refined_divisions\\{file_name_adj}.xlsx") 
        elif n == 32:
            wb = openpyxl.load_workbook("templates\\32.xlsx")    
            sheet = wb.active
            sheet['K6'] = f"Division: {file_name_adj}"
            
            sheet['B3'] = division_list[0][0]
            sheet['B4'] = division_list[0][1]
            
            sheet['B7'] = division_list[1][0]
            sheet['B8'] = division_list[1][1]
            
            sheet['B11'] = division_list[2][0]
            sheet['B12'] = division_list[2][1]
            
            sheet['B15'] = division_list[3][0]
            sheet['B16'] = division_list[3][1]
            
            sheet['B19'] = division_list[4][0]
            sheet['B20'] = division_list[4][1]
            
            sheet['B23'] = division_list[5][0]
            sheet['B24'] = division_list[5][1]
            
            sheet['B27'] = division_list[6][0]
            sheet['B28'] = division_list[6][1]
            
            sheet['B31'] = division_list[7][0]
            sheet['B32'] = division_list[7][1]
            
            sheet['B35'] = division_list[8][0]
            sheet['B36'] = division_list[8][1]
            
            sheet['B39'] = division_list[9][0]
            sheet['B40'] = division_list[9][1]
            
            sheet['B43'] = division_list[10][0]
            sheet['B44'] = division_list[10][1]
            
            sheet['B47'] = division_list[11][0]
            sheet['B48'] = division_list[11][1]
            
            sheet['B51'] = division_list[12][0]
            sheet['B52'] = division_list[12][1]
            
            sheet['B55'] = division_list[13][0]
            sheet['B56'] = division_list[13][1]
            
            sheet['B59'] = division_list[14][0]
            sheet['B60'] = division_list[14][1]
            
            sheet['B63'] = division_list[15][0]
            sheet['B64'] = division_list[15][1]

            #### halfway
            
            sheet['B67'] = division_list[16][0]
            sheet['B68'] = division_list[16][1]
            
            sheet['B71'] = division_list[17][0]
            sheet['B72'] = division_list[17][1]
            
            sheet['B75'] = division_list[18][0]
            sheet['B76'] = division_list[18][1]
            
            sheet['B79'] = division_list[19][0]
            sheet['B80'] = division_list[19][1]
            
            sheet['B83'] = division_list[20][0]
            sheet['B84'] = division_list[20][1]
            
            sheet['B87'] = division_list[21][0]
            sheet['B88'] = division_list[21][1]
            
            sheet['B91'] = division_list[22][0]
            sheet['B92'] = division_list[22][1]
            
            sheet['B95'] = division_list[23][0]
            sheet['B96'] = division_list[23][1]
            
            sheet['B99'] = division_list[24][0]
            sheet['B100'] = division_list[24][1]
            
            sheet['B103'] = division_list[25][0]
            sheet['B104'] = division_list[25][1]
            
            sheet['B107'] = division_list[26][0]
            sheet['B108'] = division_list[26][1]
            
            sheet['B111'] = division_list[27][0]
            sheet['B112'] = division_list[27][1]
            
            sheet['B115'] = division_list[28][0]
            sheet['B116'] = division_list[28][1]
            
            sheet['B119'] = division_list[29][0]
            sheet['B120'] = division_list[29][1]
            
            sheet['B123'] = division_list[30][0]
            sheet['B124'] = division_list[30][1]
            
            sheet['B127'] = division_list[31][0]
            sheet['B128'] = division_list[31][1]

            wb.save(filename = f"refined_divisions\\{file_name_adj}.xlsx") 
